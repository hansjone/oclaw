from __future__ import annotations

import io

from openpyxl import load_workbook

from runtime.chat.tool_runtime import _attachments_from_tool_result, _ref_type_for_mime
from runtime.tools.public.write_xlsx_tool import XLSX_MIME, write_xlsx_tool
from svc.files.attachment_assets import AttachmentAssetStore


def test_attachment_store_keeps_xlsx_extension(tmp_path) -> None:
    store = AttachmentAssetStore(root_dir=tmp_path / "att")
    meta = store.save_bytes(b"PK\x03\x04fake", filename="report.xlsx", mime=XLSX_MIME)
    assert meta.mime == XLSX_MIME
    path = store.get_local_path(meta.attachment_id)
    assert path is not None
    assert path.suffix == ".xlsx"
    blob, loaded = store.load_bytes(meta.attachment_id)
    assert blob.startswith(b"PK")
    assert loaded is not None
    assert loaded.name == "report.xlsx"


def test_write_xlsx_returns_attachment_id_without_deliverable(tmp_path, monkeypatch) -> None:
    store = AttachmentAssetStore(root_dir=tmp_path / "att")
    monkeypatch.setattr(
        "runtime.tools.public.write_xlsx_tool.AttachmentAssetStore",
        lambda root_dir=None: store if root_dir is None else AttachmentAssetStore(root_dir=root_dir),
    )

    spec = write_xlsx_tool()
    out = spec.handler(
        {
            "name": "alarm_summary.xlsx",
            "sheets": [
                {
                    "name": "按网元",
                    "headers": ["host_name", "severity", "count"],
                    "rows": [["NE-A", "critical", 12], ["NE-B", "major", 5]],
                },
                {
                    "name": "汇总",
                    "headers": ["metric", "value"],
                    "rows": [["total", 17]],
                },
            ],
        }
    )
    assert out.get("ok") is True
    assert out.get("attachment_id")
    assert out.get("name") == "alarm_summary.xlsx"
    assert out.get("mime") == XLSX_MIME
    assert out.get("deliverable") is not True
    assert out.get("sheet_count") == 2
    assert "hint" in out

    blob, meta = store.load_bytes(str(out["attachment_id"]))
    assert meta is not None
    assert blob[:2] == b"PK"
    wb = load_workbook(io.BytesIO(blob))
    assert wb.sheetnames == ["按网元", "汇总"]
    ws = wb["按网元"]
    assert [c.value for c in ws[1]] == ["host_name", "severity", "count"]
    assert ws["A2"].value == "NE-A"
    assert ws["C2"].value == 12
    assert ws.freeze_panes == "A2"


def test_write_xlsx_optional_workspace_path(tmp_path, monkeypatch) -> None:
    store = AttachmentAssetStore(root_dir=tmp_path / "att")
    monkeypatch.setattr(
        "runtime.tools.public.write_xlsx_tool.AttachmentAssetStore",
        lambda root_dir=None: store if root_dir is None else AttachmentAssetStore(root_dir=root_dir),
    )
    target = tmp_path / "workspace" / "tmp" / "out.xlsx"
    monkeypatch.setattr(
        "runtime.tools.public.write_xlsx_tool.resolve_workspace_path",
        lambda raw: target,
    )

    spec = write_xlsx_tool()
    out = spec.handler(
        {
            "path": "data/workspace/tmp/out.xlsx",
            "sheets": [{"name": "S1", "headers": ["a"], "rows": [[1]]}],
        }
    )
    assert out.get("ok") is True
    assert out.get("attachment_id")
    assert out.get("path") == str(target)
    assert target.exists()
    assert target.read_bytes()[:2] == b"PK"


def test_write_xlsx_requires_sheets() -> None:
    spec = write_xlsx_tool()
    out = spec.handler({})
    assert out.get("ok") is False
    assert out.get("error") == "sheets_required"


def test_write_xlsx_tool_result_maps_to_binary_ref() -> None:
    assert _ref_type_for_mime(XLSX_MIME) == "binary_ref"
    refs = _attachments_from_tool_result(
        {
            "ok": True,
            "attachment_id": "a" * 64,
            "name": "r.xlsx",
            "mime": XLSX_MIME,
            "bytes": 12,
        }
    )
    assert len(refs) == 1
    assert refs[0]["type"] == "binary_ref"
    assert refs[0].get("deliverable") is not True


def test_save_deliverable_marks_write_xlsx_attachment(tmp_path, monkeypatch) -> None:
    from runtime.tools.public.save_deliverable_attachment_tool import save_deliverable_attachment_tool

    store = AttachmentAssetStore(root_dir=tmp_path / "att")
    monkeypatch.setattr(
        "runtime.tools.public.write_xlsx_tool.AttachmentAssetStore",
        lambda root_dir=None: store if root_dir is None else AttachmentAssetStore(root_dir=root_dir),
    )
    monkeypatch.setattr(
        "runtime.tools.public.save_deliverable_attachment_tool.AttachmentAssetStore",
        lambda root_dir=None: store if root_dir is None else AttachmentAssetStore(root_dir=root_dir),
    )

    gen = write_xlsx_tool().handler(
        {"sheets": [{"headers": ["x"], "rows": [["你好"]]}], "name": "cn.xlsx"}
    )
    aid = str(gen["attachment_id"])
    marked = save_deliverable_attachment_tool().handler({"attachment_id": aid})
    assert marked.get("ok") is True
    assert marked.get("deliverable") is True
    assert marked.get("attachment_id") == aid
    assert marked.get("mime") == XLSX_MIME
