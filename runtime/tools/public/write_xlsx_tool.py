from __future__ import annotations

import io
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from runtime.tools.base import ToolSpec
from runtime.tools.path_guard import resolve_workspace_path
from svc.files.attachment_assets import AttachmentAssetStore

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_MAX_SHEETS = 50
_MAX_COLS = 200
_MAX_ROWS = 100_000
_MAX_CELL_CHARS = 32767
_MAX_AUTO_WIDTH = 48
_SHEET_NAME_RE = re.compile(r"[\[\]\*\?/\\:]")


def _safe_sheet_name(raw: str, *, index: int, used: set[str]) -> str:
    name = _SHEET_NAME_RE.sub("_", str(raw or "").strip()) or f"Sheet{index}"
    name = name[:31]
    base = name
    n = 2
    while name.lower() in used:
        suffix = f"_{n}"
        name = (base[: max(1, 31 - len(suffix))] + suffix)[:31]
        n += 1
    used.add(name.lower())
    return name


def _cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    text = str(value)
    if len(text) > _MAX_CELL_CHARS:
        return text[:_MAX_CELL_CHARS]
    return text


def _normalize_headers(raw: Any, *, col_count: int) -> list[str]:
    if isinstance(raw, list) and raw:
        headers = [str(h) if h is not None else "" for h in raw[:_MAX_COLS]]
        while len(headers) < col_count:
            headers.append(f"col_{len(headers) + 1}")
        return headers[: max(col_count, len(headers))][:_MAX_COLS]
    return [f"col_{i + 1}" for i in range(max(1, col_count))]


def _normalize_rows(raw: Any) -> tuple[list[list[Any]], int]:
    if not isinstance(raw, list):
        return [], 0
    rows: list[list[Any]] = []
    max_cols = 0
    for item in raw[:_MAX_ROWS]:
        if isinstance(item, dict):
            # Stable insertion order for dict rows without headers mapping.
            values = list(item.values())
        elif isinstance(item, (list, tuple)):
            values = list(item)
        else:
            values = [item]
        values = [_cell_value(v) for v in values[:_MAX_COLS]]
        max_cols = max(max_cols, len(values))
        rows.append(values)
    return rows, max_cols


def _apply_auto_width(ws: Any, *, col_count: int, sample_rows: list[list[Any]], headers: list[str]) -> None:
    widths: list[int] = []
    for idx in range(col_count):
        best = len(str(headers[idx])) if idx < len(headers) else 0
        for row in sample_rows[:200]:
            if idx < len(row) and row[idx] is not None:
                best = max(best, min(_MAX_AUTO_WIDTH, len(str(row[idx]))))
        widths.append(min(_MAX_AUTO_WIDTH, max(8, best + 2)))
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _build_workbook(sheets: list[dict[str, Any]], *, freeze_header: bool, auto_width: bool) -> tuple[bytes, list[dict[str, Any]]]:
    wb = Workbook()
    # Remove the default sheet; recreate from input for predictable naming.
    default = wb.active
    wb.remove(default)

    used_names: set[str] = set()
    summary: list[dict[str, Any]] = []
    for i, sheet in enumerate(sheets, start=1):
        if not isinstance(sheet, dict):
            continue
        rows, inferred_cols = _normalize_rows(sheet.get("rows"))
        headers_raw = sheet.get("headers")
        if isinstance(headers_raw, list) and headers_raw:
            col_count = max(len(headers_raw), inferred_cols, 1)
        else:
            col_count = max(inferred_cols, 1)
        headers = _normalize_headers(headers_raw, col_count=col_count)
        col_count = min(_MAX_COLS, max(len(headers), col_count))
        headers = headers[:col_count]

        title = _safe_sheet_name(str(sheet.get("name") or ""), index=i, used=used_names)
        ws = wb.create_sheet(title=title)
        ws.append(headers)
        for row in rows:
            padded = list(row[:col_count]) + [None] * max(0, col_count - len(row))
            ws.append(padded)
        if freeze_header:
            ws.freeze_panes = "A2"
        if auto_width:
            _apply_auto_width(ws, col_count=col_count, sample_rows=rows, headers=headers)
        summary.append({"name": title, "columns": col_count, "rows": len(rows)})

    if not summary:
        ws = wb.create_sheet(title="Sheet1")
        ws.append(["col_1"])
        summary.append({"name": "Sheet1", "columns": 1, "rows": 0})

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), summary


def write_xlsx_tool() -> ToolSpec:
    def _handler(args: dict[str, Any]) -> dict[str, Any]:
        sheets_raw = args.get("sheets")
        if not isinstance(sheets_raw, list) or not sheets_raw:
            return {"ok": False, "error": "sheets_required"}
        if len(sheets_raw) > _MAX_SHEETS:
            return {"ok": False, "error": "too_many_sheets", "max_sheets": _MAX_SHEETS}

        filename = str(args.get("name") or "").strip() or "report.xlsx"
        if not filename.lower().endswith(".xlsx"):
            filename = f"{filename}.xlsx"
        freeze_header = args.get("freeze_header") is not False
        auto_width = args.get("auto_width") is not False
        raw_path = str(args.get("path") or "").strip().strip('"').strip("'")

        try:
            blob, summary = _build_workbook(
                [s for s in sheets_raw if isinstance(s, dict)],
                freeze_header=freeze_header,
                auto_width=auto_width,
            )
        except Exception as exc:
            return {"ok": False, "error": "xlsx_build_failed", "detail": str(exc)}

        path_written = ""
        if raw_path:
            try:
                p = resolve_workspace_path(raw_path)
            except ValueError as exc:
                return {"ok": False, "error": str(exc)}
            if not str(p).lower().endswith(".xlsx"):
                p = p.with_suffix(".xlsx")
            try:
                p.parent.mkdir(parents=True, exist_ok=True)
                p.write_bytes(blob)
                path_written = str(p)
            except Exception as exc:
                return {"ok": False, "error": "path_write_failed", "detail": str(exc)}

        meta = AttachmentAssetStore().save_bytes(blob, filename=filename, mime=XLSX_MIME)
        out: dict[str, Any] = {
            "ok": True,
            "attachment_id": meta.attachment_id,
            "name": meta.name,
            "mime": meta.mime,
            "bytes": meta.bytes,
            "sheet_count": len(summary),
            "sheets": summary,
            "hint": (
                "Excel saved to attachment store. Not sent to channel yet. "
                "If the user asked to receive the file, call save_deliverable_attachment "
                "with this attachment_id."
            ),
        }
        if path_written:
            out["path"] = path_written
        return out

    return ToolSpec(
        name="write_xlsx",
        description=(
            "Build a real .xlsx workbook from structured sheet data (headers + rows) and save it "
            "to the attachment store. Returns attachment_id/name/mime/bytes — does NOT mark "
            "deliverable and does NOT send to WhatsApp/WeChat. To deliver, call "
            "save_deliverable_attachment(attachment_id=...). Prefer this over run_command/openpyxl."
        ),
        parameters={
            "type": "object",
            "properties": {
                "sheets": {
                    "type": "array",
                    "description": "One or more sheets. Each item: {name, headers[], rows[][]}.",
                    "items": {
                        "type": "object",
                        "properties": {
                            "name": {"type": "string", "description": "Sheet tab name (max 31 chars)."},
                            "headers": {
                                "type": "array",
                                "items": {"type": "string"},
                                "description": "Column headers (first row).",
                            },
                            "rows": {
                                "type": "array",
                                "description": "Data rows: each row is an array of cell values (string/number/bool/null).",
                                "items": {"type": "array"},
                            },
                        },
                        "required": ["rows"],
                        "additionalProperties": False,
                    },
                },
                "name": {
                    "type": "string",
                    "description": "Download filename, e.g. alarm_summary.xlsx",
                },
                "path": {
                    "type": "string",
                    "description": "Optional workspace path to also write the .xlsx file (mirror).",
                },
                "freeze_header": {
                    "type": "boolean",
                    "default": True,
                    "description": "Freeze the header row (default true).",
                },
                "auto_width": {
                    "type": "boolean",
                    "default": True,
                    "description": "Best-effort column width from sample cells (default true).",
                },
            },
            "required": ["sheets"],
            "additionalProperties": False,
        },
        handler=_handler,
        tags=frozenset({"public", "workspace", "attachment", "xlsx"}),
        read_only=False,
        risk_level="low",
    )


__all__ = ["write_xlsx_tool"]
